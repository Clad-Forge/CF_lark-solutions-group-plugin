#!/usr/bin/env python3
# ---------------------------------------------------------------------
# SOURCE OF TRUTH - edit here. This file lives in the lsg plugin repo
# (Clad-Forge/CF_lark-solutions-group-plugin); ship changes with
# ./scripts/release.sh, which bumps the version the client updates to.
#
# It began as a copy of 02-Tool/finalize_receipts.py in the LSG-005 project
# folder. That is history now: do not run 05-Plugin/build_plugin.py against
# this repo, and never copy an LSG-005 version over this file - either would
# silently discard work already released to the client.
# ---------------------------------------------------------------------
"""
LSG-005 Step 4, part two: act on the receipt check.

The skill decides - which receipt backs which expense, what the vendor is
called, what is missing. This does the mechanical half, because filing 26 files
and annotating a workbook should be exact and repeatable rather than done by
hand 26 times.

Reads  04-Receipt Check/_matches.json   the skill's judgement
       03-WR Invoice Tracker/*.xlsx     the parse output, untouched

Writes 05-Final Output/Receipts/<Category>/<renamed>   only folders that get a file
       05-Final Output/<tracker>.xlsx                  annotated copy
       04-Receipt Check/Receipt-Check-<project>.md     the run log

Why the tracker is copied and annotated rather than edited in place: stage 03
stays the parser's output and stage 05 becomes the reconciled deliverable, so
the difference between the two files is exactly what the receipt check found.

_matches.json schema
--------------------
{
  "tracker": "Voucher-Summary-NSW 2607-A-v01.xlsx",   optional, defaults to newest
  "receipts": [
    {"source": "<absolute path to the file in 01-Inbox>",
     "person": "Baptiste, Yvonne",          folder name, matches the tracker row
     "initial_surname": "Y. Baptiste",      from _extract.json
     "vendor": "Comfort Inn",
     "date": "2026-07-06",                  transaction date ON the receipt
     "category": "Lodging",                 exactly one category folder name
     "amount": 528.40,
     "status": "matched" | "mismatched" | "orphan",
     "voucher_amount": 528.40,              what the voucher says, if matched
     "note": ""}
  ],
  "unbacked": [
    {"person": "Baptiste, Yvonne", "category": "Parking, Tolls & Taxis",
     "amount": 22.10, "note": "no receipt found"}
  ]
}
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import shutil
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tracker_sheet import (BORDER, HDR_FILL, HDR_FONT, TOTAL_FILL, WARN,
                           long_path, stage_dir)

VARIANCE_FILL = "FFF8CBAD"
OK_FILL = "FFE7EDEB"
ILLEGAL = r'[\\/:*?"<>|]'


def clean(text, fallback="UNKNOWN"):
    """Make one filename field safe. Never silently empties a field."""
    text = re.sub(ILLEGAL, "-", str(text or "")).strip()
    text = re.sub(r"\s+-\s+", " ", text)      # keep " - " as the field separator
    text = re.sub(r"\s{2,}", " ", text).strip(" .")
    return text or fallback


def receipt_name(rec):
    """YYYY.MM.DD - Vendor - Category - F. Lastname.ext"""
    source = Path(rec["source"])
    date = str(rec.get("date") or "").strip()
    try:
        stamp = dt.date.fromisoformat(date).strftime("%Y.%m.%d")
    except ValueError:
        stamp = "0000.00.00"                  # unreadable date, deliberately obvious
    parts = [stamp, clean(rec.get("vendor"), "UNKNOWN VENDOR"),
             clean(rec.get("category"), "MISC"), clean(rec.get("initial_surname"), "UNKNOWN")]
    return " - ".join(parts) + source.suffix.lower()


def exists(path):
    return os.path.exists(long_path(path))


def unique(folder, name):
    """' (2)', ' (3)' before the extension. Two same-day charges is normal."""
    target = folder / name
    if not exists(target):
        return target
    stem, suffix = target.stem, target.suffix
    for n in range(2, 100):
        candidate = folder / f"{stem} ({n}){suffix}"
        if not exists(candidate):
            return candidate
    raise RuntimeError(f"cannot find a free name for {name}")


def file_receipts(receipts, final):
    """Copy each receipt to its category folder under its new name.

    Folders are created HERE, one at a time, as a file lands in each - never up
    front. A folder existing therefore means it has something in it."""
    filed, failed = [], []
    for rec in receipts:
        source = Path(rec["source"])
        if not source.is_file():
            failed.append((rec, "source file not found"))
            continue
        folder = final / "Receipts" / clean(rec.get("category"), "MISC")
        os.makedirs(long_path(folder), exist_ok=True)
        target = unique(folder, receipt_name(rec))
        try:
            shutil.copy2(long_path(source), long_path(target))   # copy, never move
        except OSError as exc:
            failed.append((rec, str(exc)))
            continue
        filed.append({**rec, "filed_as": target.name,
                      "category_folder": folder.name})
    return filed, failed


def annotate_tracker(src, dest, filed, unbacked):
    """Copy the tracker into 05 and add two receipt columns to the copy.

    Only appends - no existing cell, formula or number is touched. Stage 03
    keeps the parser's output exactly as produced."""
    by_person = defaultdict(list)
    for rec in filed:
        by_person[rec["person"]].append(rec)
    gaps = defaultdict(list)
    for gap in unbacked:
        gaps[gap["person"]].append(gap)

    wb = load_workbook(src)
    ws = wb["Invoice Tracker"]
    header_row, first = 5, 6
    last_col = ws.max_column
    while last_col > 1 and ws.cell(header_row, last_col).value is None:
        last_col -= 1

    cols = {"Receipt Check": last_col + 1, "Receipts Missing": last_col + 2}
    for label, idx in cols.items():
        cell = ws.cell(header_row, idx, label)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.column_dimensions[ws.cell(header_row, cols["Receipt Check"]).column_letter].width = 14
    ws.column_dimensions[ws.cell(header_row, cols["Receipts Missing"]).column_letter].width = 52

    row = first
    while ws.cell(row, 1).value not in (None, "TOTALS"):
        person = str(ws.cell(row, 1).value or "")
        # the tracker carries the employee name; _matches carries the folder name
        mine = next((v for k, v in by_person.items() if k.split(",")[0].strip().lower()
                     in person.lower()), [])
        my_gaps = next((v for k, v in gaps.items() if k.split(",")[0].strip().lower()
                        in person.lower()), [])
        bad = [r for r in mine if r.get("status") in ("mismatched", "orphan")]

        if my_gaps or bad:
            verdict = "GAPS"
        elif mine:
            verdict = "OK"
        else:
            verdict = "NO RECEIPTS"

        detail = "; ".join(
            [f"{g['category']} {float(g['amount']):,.2f} unbacked" for g in my_gaps] +
            [f"{r['category']} receipt {float(r.get('amount') or 0):,.2f} vs voucher "
             f"{float(r.get('voucher_amount') or 0):,.2f}"
             for r in bad if r.get("status") == "mismatched"] +
            [f"{r['vendor']} {float(r.get('amount') or 0):,.2f} matches nothing"
             for r in bad if r.get("status") == "orphan"])

        check = ws.cell(row, cols["Receipt Check"], verdict)
        check.border = BORDER
        check.fill = PatternFill("solid", start_color=(OK_FILL if verdict == "OK" else VARIANCE_FILL))
        check.font = Font(bold=True, size=10)
        note = ws.cell(row, cols["Receipts Missing"], detail)
        note.border, note.alignment = BORDER, Alignment(wrap_text=True, vertical="top")
        if detail:
            note.fill = PatternFill("solid", start_color=WARN)
        row += 1

    if ws.cell(row, 1).value == "TOTALS":
        for idx in cols.values():
            ws.cell(row, idx).fill = TOTAL_FILL
            ws.cell(row, idx).border = BORDER
        ws.cell(row, cols["Receipt Check"],
                f"{sum(1 for r in filed)} filed").font = Font(bold=True, size=10)

    wb.save(dest)
    wb.close()
    return cols


def write_report(path, project, tracker_name, filed, failed, unbacked, receipts):
    counts = defaultdict(int)
    for rec in receipts:
        counts[rec.get("status", "unknown")] += 1
    lines = [
        f"# Receipt check — {project}", "",
        f"Run {dt.date.today().isoformat()} against `{tracker_name}`.", "",
        "This is the run log. **The workbook in `05-Final Output` is the record** —",
        "it carries the same findings as `Receipt Check` and `Receipts Missing`",
        "columns, so they travel with the deliverable.", "",
        "| | |", "|---|---|",
        f"| Receipts filed | {len(filed)} |",
        f"| Matched | {counts['matched']} |",
        f"| Mismatched | {counts['mismatched']} |",
        f"| Orphan | {counts['orphan']} |",
        f"| Expenses with no receipt | {len(unbacked)} |",
        f"| Could not be filed | {len(failed)} |", "",
    ]

    if unbacked:
        lines += ["## Unbacked expenses", "",
                  "Claimed on a voucher with no receipt behind it. **This is the list that",
                  "matters** — it is what gets asked about before West Run invoices CDI.", "",
                  "| Person | Category | Amount | Note |", "|---|---|---|---|"]
        for g in sorted(unbacked, key=lambda x: (x["person"], x["category"])):
            lines.append(f"| {g['person']} | {g['category']} | "
                         f"{float(g['amount']):,.2f} | {g.get('note', '')} |")
        lines.append("")

    bad = [r for r in receipts if r.get("status") in ("mismatched", "orphan")]
    if bad:
        lines += ["## Receipts that do not tie", "",
                  "| Person | Receipt | Category | Receipt says | Voucher says | |",
                  "|---|---|---|---|---|---|"]
        for r in sorted(bad, key=lambda x: (x["person"], x.get("vendor", ""))):
            lines.append(
                f"| {r['person']} | {r.get('vendor', '?')} | {r.get('category', '?')} | "
                f"{float(r.get('amount') or 0):,.2f} | "
                f"{float(r.get('voucher_amount') or 0):,.2f} | {r['status']} |")
        lines.append("")

    unknown = [r for r in filed if "UNKNOWN" in r.get("filed_as", "")]
    if unknown:
        lines += ["## Needs a human to name", "",
                  "Filed with a placeholder rather than a guess.", ""]
        lines += [f"- `{r['filed_as']}` — from `{Path(r['source']).name}`" for r in unknown]
        lines.append("")

    if failed:
        lines += ["## Could not be filed", ""]
        lines += [f"- `{Path(r['source']).name}` — {why}" for r, why in failed]
        lines.append("")

    if filed:
        lines += ["## Filed", ""]
        for folder in sorted({r["category_folder"] for r in filed}):
            lines.append(f"**{folder}**")
            lines += [f"- `{r['filed_as']}`" for r in sorted(filed, key=lambda x: x["filed_as"])
                      if r["category_folder"] == folder]
            lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")


def main(argv=None):
    ap = argparse.ArgumentParser(description="Act on the Step 4 receipt check.")
    ap.add_argument("--project", required=True)
    ap.add_argument("--matches", help="defaults to <project>/04-*/_matches.json")
    args = ap.parse_args(argv)

    project = Path(args.project).resolve()
    check_dir = stage_dir(project, "04", "04-Receipt Check")
    final = stage_dir(project, "05", "05-Final Output")
    tracker_dir = stage_dir(project, "03", "03-WR Invoice Tracker")

    matches_path = Path(args.matches) if args.matches else check_dir / "_matches.json"
    if not matches_path.is_file():
        print(f"no _matches.json at {matches_path} - the skill writes it before "
              f"calling this", file=sys.stderr)
        return 2
    data = json.loads(matches_path.read_text(encoding="utf-8"))
    receipts = data.get("receipts", [])
    unbacked = data.get("unbacked", [])

    # Newest by modification time, not by name. Name sorting looked fine until
    # the workbooks were renamed: "… - NSW 2607-A (2).xlsx" sorts BEFORE
    # "… - NSW 2607-A.xlsx" because a space precedes a period, so a same-day
    # re-run would be silently ignored in favour of the earlier file.
    named = data.get("tracker")
    candidates = sorted((p for p in tracker_dir.glob("*.xlsx") if not p.name.startswith("~$")),
                        key=lambda p: p.stat().st_mtime)
    src = (tracker_dir / named) if named else (candidates[-1] if candidates else None)
    if not src or not src.is_file():
        print(f"no tracker workbook in {tracker_dir} - run the parser first",
              file=sys.stderr)
        return 2

    final.mkdir(parents=True, exist_ok=True)
    filed, failed = file_receipts(receipts, final)
    annotate_tracker(src, final / src.name, filed, unbacked)
    report = check_dir / f"Receipt-Check-{project.name}.md"
    write_report(report, project.name, src.name, filed, failed, unbacked, receipts)

    print(f"filed:   {len(filed)} receipt(s) into "
          f"{len({r['category_folder'] for r in filed})} category folder(s)")
    if failed:
        print(f"  !! {len(failed)} could not be filed")
    print(f"tracker: {final / src.name}")
    print(f"         + Receipt Check / Receipts Missing columns")
    print(f"report:  {report}")
    if unbacked:
        print(f"  {len(unbacked)} expense(s) with no receipt:")
        for g in unbacked:
            print(f"    {g['person']:<20} {g['category']:<24} {float(g['amount']):>9,.2f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
