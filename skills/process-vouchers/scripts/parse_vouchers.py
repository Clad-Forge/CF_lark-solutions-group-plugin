#!/usr/bin/env python3
# ---------------------------------------------------------------------
# COPY - do not edit here. The original is 02-Tool/parse_vouchers.py in the LSG-005
# project folder; this file is overwritten by 05-Plugin/build_plugin.py.
# ---------------------------------------------------------------------
"""
LSG-005 voucher parser.

Reads every West Run, LLC travel voucher in an inbox folder and produces ONE
workbook: a Master summary sheet up front, then one sheet per voucher.

The kept field set is unchanged from the first build:
  A9   employee name
  I9   voucher date
  rows 17-38, columns A, C, D, E, F, G, H, I, J
       date, POV miles, mileage $, air fare, car rental, rental fuel,
       hotel, per diem/meals, parking/taxi/tolls
  D39:J39  column totals as the voucher reports them
  due employee - located by LABEL, not by a fixed cell (see below)

Two voucher forms are in circulation and both parse:

  West Run "Travel Voucher" (current - 01-Reference/Dummy Voucher (Actual).xlsx)
      Header block and the row 17-38 grid are identical to the older form.
      Everything below row 39 differs: a Section B independent-contractor
      labor block sits at rows 40-41, and the totals moved out of column J into
      column I and down three rows -
          I42 Section A total travel  ·  I43 total trip expenses
          I44 cash advance            ·  I45 total reimbursement/labor due
  Older "EMPLOYEE TRAVEL / EXPENSE VOUCHER" form
          J40 total trip expenses · J41 cash advance · J42 due employee

Rather than carry two cell maps, the footer is resolved by reading the labels
in rows 40-53 and taking the rightmost number on the matching row. That also
survives the row drift the Numbers-to-Excel export introduces.

Column B (Travel From/To) is deliberately not carried through. It is still read
internally to decide whether an expense row is in use.

The SSN/EIN in F9 on the current form is never read and must never be added -
it is the one field on the voucher that would turn a summary workbook into a
PII liability.

No AI. Fixed cell addresses, same answer every run. Source files are opened
read-only and never modified.

Usage:
    python parse_vouchers.py --project "04-Projects/NSW 2607-A"
    python parse_vouchers.py --inbox <folder> --out <file.xlsx> [--exceptions <folder>]
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Needed for the bundled interpreter. An embeddable CPython takes its whole
# sys.path from python313._pth and does NOT add the running script's own folder,
# so a plain sibling import fails there while working fine everywhere else.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from tracker_sheet import (safe_field, scaffold_final_output, split_name,  # noqa: E402
                           stage_dir, tracker_name, write_tracker)

# ---------------------------------------------------------------- cell layout

SHEET_NAME = "Voucher"
CELL_NAME = "A9"
CELL_DATE = "I9"
ROW_FIRST, ROW_LAST = 17, 38
TOTALS_ROW = 39
COL_ROUTE = 2  # read for row occupancy only, not carried through
COL_LAST = 10  # column J

# The footer below the grid is the only part of the voucher whose geometry
# differs between the two forms, so it is found by label instead of by cell.
# Rows 40-53 covers the West Run footer (which runs to row 45 plus signature
# rows) with room for the export to shift things a little.
FOOTER_ROWS = (TOTALS_ROW + 1, TOTALS_ROW + 14)

# key -> alternative label patterns; a pattern matches when EVERY one of its
# fragments appears in the normalised cell text.
FOOTER_LABELS = {
    "travel_total":  (("total travel",),),
    "trip_expenses": (("total trip expenses",),),
    "advance":       (("cash advance",),),
    "due":           (("due employee",), ("reimbursement", "due")),
}

# What the SSN/EIN field looks like, so the parser can assert it never leaks
# into the output. Read the layout note before touching this.
CELL_SSN = "F9"

# key, source column, heading, format  -- the yellow grid columns, in order
LINE_COLUMNS = [
    ("date",       1, "Date",               "date"),
    ("pov_miles",  3, "POV Miles",          "miles"),
    ("mileage",    4, "Mileage $",          "money"),
    ("airfare",    5, "Air Fare",           "money"),
    ("car_rental", 6, "Car Rental",         "money"),
    ("fuel",       7, "Rental Fuel",        "money"),
    ("hotel",      8, "Hotel",              "money"),
    ("per_diem",   9, "Per Diem / Meals",   "money"),
    ("taxi",      10, "Taxi/Tolls/Parking", "money"),
]

# the money categories that roll up to the master table
CATEGORIES = ["mileage", "airfare", "car_rental", "fuel", "hotel", "per_diem", "taxi"]
CATEGORY_LABELS = {
    "mileage": "Mileage $",
    "airfare": "Air Fare",
    "car_rental": "Car Rental",
    "fuel": "Rental Fuel",
    "hotel": "Hotel",
    "per_diem": "Per Diem / Meals",
    "taxi": "Taxi/Tolls/Parking",
}

DATE_FORMATS = [
    "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y",
    "%B %d %Y", "%B %d, %Y", "%b %d %Y", "%b %d, %Y",
    "%d-%b-%Y", "%d %B %Y",
]

PENNY = 0.005  # tolerance when comparing our line sums to the voucher's totals

# Mileage rate the contractors are supposed to bill at. Override with --rate.
# A voucher can carry three different rates in one trip (the sample uses 0.725,
# 0.70 and 0.67), so there is no safe way to infer it - it has to be told to us.
# 0.725 is what West Run's own tracker bills at as of 2026-08-09. It is a
# default, not a constant: GSA changes it, and the launcher asks every run.
DEFAULT_RATE = 0.725

# audit columns appended to each voucher sheet, after the kept fields
AUDIT_COLUMNS = [
    ("rate_used",   "Rate Used",      "rate"),
    ("rate_should", "Expected $",     "money"),
    ("rate_diff",   "Diff",           "money"),
    ("formula",     "Source Formula", None),
    ("formula_note", "Formula Check", None),
]

RE_REF = re.compile(r"\bC(\d+)\b", re.IGNORECASE)
RE_RATE = re.compile(r"\*\s*([0-9]*\.?[0-9]+)")

# -------------------------------------------------------------------- styling

FMT = {
    "money": '#,##0.00',
    "miles": '#,##0',
    "date": 'yyyy-mm-dd',
    "rate": '0.0000',
}
HDR_FILL = PatternFill("solid", start_color="FF1F3864")
HDR_FONT = Font(bold=True, color="FFFFFFFF", size=10)
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", start_color="FFD9E2F3")
WARN_FILL = PatternFill("solid", start_color="FFFFF2CC")
ERR_FILL = PatternFill("solid", start_color="FFF8CBAD")
LABEL_FONT = Font(bold=True, size=10)
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ------------------------------------------------------------------- coercion

def as_number(value):
    """Return a float, or None if the cell is blank or not numeric."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace("$", "").replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def as_date(value):
    """Return (date, note). Handles real dates and the text dates contractors type."""
    if value is None:
        return None, None
    if isinstance(value, dt.datetime):
        return value.date(), None
    if isinstance(value, dt.date):
        return value, None
    text = str(value).strip()
    if not text:
        return None, None
    for fmt in DATE_FORMATS:
        try:
            return dt.datetime.strptime(text, fmt).date(), None
        except ValueError:
            continue
    return None, f"unreadable date {text!r}"


def blank(value):
    return value is None or (isinstance(value, str) and not value.strip())


# ------------------------------------------------------------- footer locator

def norm_label(value):
    """Cell text flattened for label matching: lowercase, single-spaced."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def rightmost_number(ws, row, col_from, col_to=COL_LAST):
    """The last numeric cell on a row, scanning right to left.

    The label sits in a merged block on the left and its amount in the last
    populated column - column J on the older form, column I on the West Run
    one. Scanning backwards finds it without caring which."""
    for col in range(col_to, col_from - 1, -1):
        value = as_number(ws.cell(row, col).value)
        if value is not None:
            return f"{get_column_letter(col)}{row}", value
    return None, None


def find_footer(ws):
    """Locate the totals under the grid by their labels.

    Returns {key: {"cell": "I45", "value": 1999.85, "label": "..."}} for
    whichever of travel_total / trip_expenses / advance / due are present.
    First match wins, so the certification and signature rows below cannot
    overwrite a real total."""
    found = {}
    first, last = FOOTER_ROWS
    for row in range(first, min(last, ws.max_row) + 1):
        for col in range(1, COL_LAST + 1):
            text = norm_label(ws.cell(row, col).value)
            if len(text) < 3:
                continue
            for key, patterns in FOOTER_LABELS.items():
                if key in found:
                    continue
                if any(all(frag in text for frag in pat) for pat in patterns):
                    cell, value = rightmost_number(ws, row, col + 1)
                    found[key] = {"cell": cell, "value": value, "label": text}
    return found


def date_outliers(lines, tolerance_days=45):
    """Line dates that sit far away from the rest of the voucher.

    A trip is days or weeks long, so a line months off the others is a keying
    error - the actual sample has a July voucher with one April date in it.
    Compared against the median so a single bad row cannot drag the reference."""
    dated = [(ln["row"], ln["date"]) for ln in lines if ln.get("date")]
    if len(dated) < 3:
        return []
    ordered = sorted(d for _, d in dated)
    median = ordered[len(ordered) // 2]
    return [
        (row, date, abs((date - median).days))
        for row, date in dated
        if abs((date - median).days) > tolerance_days
    ]


# split_name lives in tracker_sheet so the naming rules sit in one place;
# re-exported here because extract_receipts imports it from this module.


# What a receipt can arrive as. Emailed receipts are usually PDFs; phone photos
# of paper ones are not. HEIC is what an iPhone sends unless the sender changed
# a setting, so it has to be on the list even though nothing can read it yet.
RECEIPT_TYPES = {".pdf", ".png", ".jpg", ".jpeg", ".heic", ".heif",
                 ".tif", ".tiff", ".webp", ".bmp", ".gif"}


def usable(path):
    """Skip Excel lock files and anything hidden or OS-generated."""
    return not path.name.startswith("~$") and not path.name.startswith(".")


# The longest path this pipeline creates, measured from the PROJECT ROOT:
#
#   <project>\Output\Receipts\Parking, Tolls & Taxis\<renamed receipt>
#   \________ 38 chars _____________________________/\___ ~80 chars __/
#
# Measured from the project rather than the inbox because that is where the long
# ones actually live - an inbox path is a person folder and an as-received
# filename, which is shorter. Getting this wrong the other way is not harmless:
# a warning that cries wolf on a folder that is genuinely fine teaches whoever
# sees it to ignore the one that matters.
NESTING_ALLOWANCE = 120
MAX_PATH = 260


def warn_if_path_is_long(inbox):
    """Say so - loudly - when the folder is deep enough to lose files.

    Windows silently stops returning matches from rglob() once a path crosses
    260 characters. Not an error, not an exception: the file simply is not in
    the results. On a pipeline whose entire job is to notice missing money, a
    voucher that vanishes because the folder is nested too deep is the worst
    failure available - it would report a clean run over an incomplete set.

    The writes themselves survive, because every one goes through long_path().
    The damage is subtler: files land where Explorer and ordinary PowerShell
    cannot delete or rename them, so the output becomes something a human
    cannot tidy up by hand. Seen for real 2026-08-28 on the old five-stage
    layout, where '05-Final Output (Manually Verify)' pushed the filed receipts
    past the limit.

    A warning rather than a hard stop, because the real fix is to move the
    project folder and only a human can decide where."""
    project = Path(inbox).resolve().parent
    length = len(str(project))
    if length + NESTING_ALLOWANCE <= MAX_PATH:
        return
    print(
        f"\n  !! WARNING - this project folder is {length} characters deep.\n"
        f"     Windows starts dropping files from searches past {MAX_PATH} without\n"
        f"     reporting an error, so vouchers or receipts MAY BE MISSED SILENTLY,\n"
        f"     and filed receipts may land where Explorer cannot delete them.\n"
        f"     Check the counts below against what you expect, and move the project\n"
        f"     somewhere shallower before trusting this run.\n",
        file=sys.stderr,
    )


def collect_submissions(inbox):
    """One folder per person, holding their voucher and their receipts.

    Returns [(voucher_path_or_None, [receipt paths], folder name)].

    A folder with receipts but no voucher still comes back, so somebody who
    sent receipts and forgot the voucher shows up as a problem rather than
    silently not existing. Loose .xlsx sitting straight in the inbox are still
    picked up, so an inbox filled the old flat way keeps working."""
    warn_if_path_is_long(inbox)
    submissions = []
    for folder in sorted(p for p in inbox.iterdir() if p.is_dir() and usable(p)):
        vouchers = sorted(p for p in folder.rglob("*.xlsx") if usable(p) and p.is_file())
        receipts = sorted(p for p in folder.rglob("*")
                          if p.is_file() and usable(p)
                          and p.suffix.lower() in RECEIPT_TYPES)
        if not vouchers:
            submissions.append((None, receipts, folder.name))
            continue
        for voucher in vouchers:
            # More than one voucher in a folder is unusual but legal - a person
            # can file twice for one project. Receipts attach to each, because
            # nothing on a receipt says which voucher it belongs to.
            submissions.append((voucher, receipts, folder.name))
    for voucher in sorted(p for p in inbox.glob("*.xlsx") if usable(p)):
        submissions.append((voucher, [], None))
    return submissions


def orphan_result(receipts, folder_name):
    """A person folder holding receipts but no voucher."""
    return {
        "source": folder_name, "employee": "", "tdy": "", "exercise": "",
        "voucher_date": None, "lines": [], "reported": {}, "computed": {},
        "due_employee": None, "due_cell": None, "labor": None, "advance": None,
        "rate_issues": 0, "rates_used": [], "status": "ERROR",
        "receipts": [p.name for p in receipts], "folder": folder_name,
        "notes": [f"{len(receipts)} receipt(s) but no voucher in this folder"],
    }


def name_sort_key(result):
    """Alphabetical by surname then forename. Anything with no readable name
    sorts last rather than jumping to the top on an empty string."""
    last, first, _ = split_name(result.get("employee") or "")
    if not last:
        return (1, "", "", result["source"].lower())
    return (0, last.lower(), first.lower(), result["source"].lower())


def sheet_title(name, source, used):
    """Excel sheet name: <=31 chars, no []:*?/\\, unique within the workbook.

    'Last, First' - commas are legal in a sheet name, and the vouchers already
    arrive in that order, so the tab reads the way the roster does."""
    last, first, _ = split_name(name)
    if last:
        base = f"{last.title()}, {first.title()}" if first else last.title()
    else:
        base = Path(source).stem
    base = re.sub(r"[\[\]:*?/\\]", "-", base).strip() or "Voucher"
    base = base[:31]
    title, n = base, 2
    while title.lower() in used:
        suffix = f" ({n})"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title.lower())
    return title


# --------------------------------------------------------------------- parser

def audit_mileage(line, src_row, formula, expected_rate):
    """Work out what rate the contractor actually billed, and whether the
    original equation is sound. Returns (rate_used, note)."""
    miles, amount = line.get("pov_miles"), line.get("mileage")
    notes = []

    rate_used = None
    if miles and amount is not None and miles > 0:
        rate_used = round(amount / miles, 4)
    elif amount is not None and amount > 0:
        # Mileage billed on a row with no POV miles. On the older form this was
        # always a stale copy-paste formula pointing at another row's mileage,
        # caught below by reading the equation. The Numbers export flattens
        # every formula to a literal, so the equation is gone and this is the
        # only thing left that catches it.
        notes.append(f"{amount:.2f} mileage billed with no POV miles")

    if formula:
        ref = RE_REF.search(formula)
        if ref and int(ref.group(1)) != src_row:
            notes.append(f"formula reads C{ref.group(1)}, should be C{src_row}")
        rate_in_formula = RE_RATE.search(formula)
        if rate_in_formula:
            written = float(rate_in_formula.group(1))
            if abs(written - expected_rate) > 1e-9:
                notes.append(f"rate {written} not {expected_rate}")
    elif rate_used is not None and abs(rate_used - expected_rate) > 0.0005:
        notes.append(f"typed value implies {rate_used}, not {expected_rate}")

    return rate_used, "; ".join(notes)


def parse_voucher(path: Path, expected_rate: float = DEFAULT_RATE) -> dict:
    """Pull the highlighted fields out of one voucher file."""
    result = {
        "source": path.name,
        "status": "OK",
        "notes": [],
        "employee": "",
        "tdy": "",
        "exercise": "",
        "voucher_date": None,
        "lines": [],
        "reported": {},
        "computed": {},
        "due_employee": None,
        "due_cell": None,
        "labor": None,
        "advance": None,
        "receipts": [],
        "rate_issues": 0,
        "rates_used": [],
        "expected_rate": expected_rate,
    }

    try:
        wb = load_workbook(path, data_only=True, read_only=False)
        # second pass, formulas intact - this is what makes the rate auditable
        wbf = load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:  # unreadable / not a workbook / password
        result["status"] = "ERROR"
        result["notes"].append(f"cannot open: {exc}")
        return result

    try:
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]
        wsf = wbf[ws.title]
        if SHEET_NAME not in wb.sheetnames:
            result["notes"].append(f"no '{SHEET_NAME}' sheet; read '{ws.title}' instead")

        # --- header
        name = ws[CELL_NAME].value
        if blank(name):
            result["status"] = "ERROR"
            result["notes"].append(f"no employee name in {CELL_NAME}")
        else:
            result["employee"] = str(name).strip()

        # C13/C14 are not reimbursement fields and are not carried per line -
        # they are read once because the tracker's header block is keyed on
        # them: the exercise name in C14 is what a project folder is named.
        result["tdy"] = str(ws["C13"].value).strip() if not blank(ws["C13"].value) else ""
        result["exercise"] = str(ws["C14"].value).strip() if not blank(ws["C14"].value) else ""

        vdate, note = as_date(ws[CELL_DATE].value)
        result["voucher_date"] = vdate
        if note:
            result["notes"].append(f"{CELL_DATE}: {note}")
        elif vdate is None:
            result["notes"].append(f"no voucher date in {CELL_DATE}")

        # --- expense rows
        # Occupancy is the date or the route, never the mileage column: the
        # template leaves stale formulas in D on rows that hold no data.
        for row in range(ROW_FIRST, ROW_LAST + 1):
            date_raw = ws.cell(row, 1).value
            route_raw = ws.cell(row, COL_ROUTE).value
            if blank(date_raw) and blank(route_raw):
                continue

            line = {"row": row}
            for key, col, _, kind in LINE_COLUMNS:
                raw = ws.cell(row, col).value
                if kind == "date":
                    value, note = as_date(raw)
                    if note:
                        result["notes"].append(f"row {row}: {note}")
                else:
                    value = as_number(raw)
                line[key] = value

            # the original mileage equation, verbatim, plus what it implies
            raw_formula = wsf.cell(row, 4).value
            line["formula"] = raw_formula if isinstance(raw_formula, str) and raw_formula.startswith("=") else ""
            rate_used, note = audit_mileage(line, row, line["formula"], expected_rate)
            line["rate_used"] = rate_used
            line["formula_note"] = note or ("typed value" if not line["formula"] else "ok")
            if note:
                result["rate_issues"] += 1
            if rate_used is not None:
                result["rates_used"].append(rate_used)

            result["lines"].append(line)

        if not result["lines"]:
            result["status"] = "ERROR"
            result["notes"].append("no expense rows found")

        # --- totals as the voucher reports them
        for key, col, _, kind in LINE_COLUMNS:
            if kind == "date":
                continue
            if key == "pov_miles":
                continue  # C39 is not a total on this template
            result["reported"][key] = as_number(ws.cell(TOTALS_ROW, col).value)

        # --- totals as the lines actually add up
        for key in CATEGORIES + ["pov_miles"]:
            values = [ln[key] for ln in result["lines"] if ln.get(key) is not None]
            result["computed"][key] = round(sum(values), 2) if values else 0.0

        # --- reconcile
        for key in CATEGORIES:
            reported = result["reported"].get(key)
            computed = result["computed"].get(key, 0.0)
            if reported is None:
                continue
            if abs(reported - computed) > PENNY:
                result["notes"].append(
                    f"{CATEGORY_LABELS[key]} total says {reported:.2f}, lines add to {computed:.2f}"
                )
                if result["status"] == "OK":
                    result["status"] = "WARN"

        # --- line dates that do not belong to this trip
        for row, bad_date, gap in date_outliers(result["lines"]):
            result["notes"].append(
                f"row {row}: date {bad_date.isoformat()} is {gap} days from the rest "
                f"of this voucher - likely a keying error"
            )
            if result["status"] == "OK":
                result["status"] = "WARN"

        # --- footer, found by label so both voucher forms work
        footer = find_footer(ws)
        due = footer.get("due", {}).get("value")
        result["due_employee"] = due
        result["due_cell"] = footer.get("due", {}).get("cell")
        if due is None:
            where = "; ".join(sorted(footer)) or "nothing"
            result["notes"].append(
                f"no amount due found under the grid (matched {where})"
            )
            if result["status"] == "OK":
                result["status"] = "WARN"

        # Section B independent-contractor labor, where the form has one. It is
        # the gap between total trip expenses and Section A travel, so it needs
        # no cell of its own. Not carried as a column - the kept field set is
        # travel only - but Due will not equal the travel categories when it is
        # billed, and that has to be visible rather than look like a bad sum.
        # Cash advance. Blank means "no advance", not 0.00 - but the tracker
        # has to subtract something, so it is normalised to 0.0 there, not here.
        result["advance"] = footer.get("advance", {}).get("value")

        travel_total = footer.get("travel_total", {}).get("value")
        trip_expenses = footer.get("trip_expenses", {}).get("value")
        if travel_total is not None and trip_expenses is not None:
            labor = round(trip_expenses - travel_total, 2)
            result["labor"] = labor
            if abs(labor) > PENNY:
                result["notes"].append(
                    f"Due includes {labor:.2f} Section B labor, which is not in the "
                    f"travel categories"
                )
                if result["status"] == "OK":
                    result["status"] = "WARN"

            grid_total = round(sum(v for v in result["reported"].values() if v), 2)
            if abs(travel_total - grid_total) > PENNY:
                result["notes"].append(
                    f"Section A total travel says {travel_total:.2f}, the row {TOTALS_ROW} "
                    f"column totals add to {grid_total:.2f}"
                )
                if result["status"] == "OK":
                    result["status"] = "WARN"

        if result["rate_issues"]:
            result["notes"].append(
                f"{result['rate_issues']} mileage line(s) fail the rate check (expected {expected_rate})"
            )
            if result["status"] == "OK":
                result["status"] = "WARN"

    finally:
        wb.close()
        wbf.close()

    if result["status"] == "OK" and result["notes"]:
        result["status"] = "WARN"
    return result


# --------------------------------------------------------------------- output

def autosize(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_voucher_sheet(wb, result, title):
    ws = wb.create_sheet(title)

    ws["A1"] = "Employee"
    ws["B1"] = result["employee"] or "(missing)"
    ws["A2"] = "Voucher Date"
    ws["B2"] = result["voucher_date"]
    ws["B2"].number_format = FMT["date"]
    ws["A3"] = "Source File"
    ws["B3"] = result["source"]
    ws["A4"] = "Status"
    ws["B4"] = result["status"]
    for cell in ("A1", "A2", "A3", "A4"):
        ws[cell].font = LABEL_FONT
    if result["status"] == "WARN":
        ws["B4"].fill = WARN_FILL
    elif result["status"] == "ERROR":
        ws["B4"].fill = ERR_FILL

    rate = result.get("expected_rate", DEFAULT_RATE)
    ws["A5"] = "Rate Checked Against"
    ws["A5"].font = LABEL_FONT
    ws["B5"] = rate
    ws["B5"].number_format = FMT["rate"]

    top = 7
    headings = [h for _, _, h, _ in LINE_COLUMNS] + [h for _, h, _ in AUDIT_COLUMNS]
    for idx, heading in enumerate(headings, start=1):
        cell = ws.cell(top, idx, heading)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    n_kept = len(LINE_COLUMNS)
    col_miles, col_mileage = 2, 3          # POV Miles, Mileage $
    col_rate = n_kept + 1                  # Rate Used
    col_should, col_diff = n_kept + 2, n_kept + 3

    first_data = top + 1
    row_at = top
    for line in result["lines"]:
        row_at += 1
        for idx, (key, _, _, kind) in enumerate(LINE_COLUMNS, start=1):
            cell = ws.cell(row_at, idx, line.get(key))
            cell.number_format = FMT[kind]
            cell.border = BORDER

        m, d = get_column_letter(col_miles), get_column_letter(col_mileage)
        should = get_column_letter(col_should)

        # live equations, so editing a line re-checks itself
        cell = ws.cell(row_at, col_rate, f'=IF(N({m}{row_at})=0,"",ROUND({d}{row_at}/{m}{row_at},4))')
        cell.number_format, cell.border = FMT["rate"], BORDER

        cell = ws.cell(row_at, col_should, f'=IF(N({m}{row_at})=0,"",ROUND({m}{row_at}*$B$5,2))')
        cell.number_format, cell.border = FMT["money"], BORDER

        cell = ws.cell(row_at, col_diff, f'=IF({should}{row_at}="","",ROUND({d}{row_at}-{should}{row_at},2))')
        cell.number_format, cell.border = FMT["money"], BORDER

        cell = ws.cell(row_at, n_kept + 4, line.get("formula", ""))
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left")

        note_cell = ws.cell(row_at, n_kept + 5, line.get("formula_note", ""))
        note_cell.border = BORDER
        if line.get("formula_note") not in ("ok", "typed value", ""):
            note_cell.fill = WARN_FILL

    last_data = row_at

    row_at += 1
    label = ws.cell(row_at, 1, "TOTAL")
    label.font, label.fill, label.border = TOTAL_FONT, TOTAL_FILL, BORDER
    for idx, (key, _, _, kind) in enumerate(LINE_COLUMNS, start=1):
        if kind == "date":
            continue
        col = get_column_letter(idx)
        cell = ws.cell(row_at, idx, f"=SUM({col}{first_data}:{col}{last_data})")
        cell.number_format, cell.font, cell.fill, cell.border = FMT[kind], TOTAL_FONT, TOTAL_FILL, BORDER
    for idx in (col_should, col_diff):
        col = get_column_letter(idx)
        cell = ws.cell(row_at, idx, f"=SUM({col}{first_data}:{col}{last_data})")
        cell.number_format, cell.font, cell.fill, cell.border = FMT["money"], TOTAL_FONT, TOTAL_FILL, BORDER
    total_row = row_at

    row_at += 2
    ws.cell(row_at, 1, "Due Employee").font = LABEL_FONT
    due = ws.cell(row_at, 3, result["due_employee"])
    due.number_format = '"$"#,##0.00'
    due.font = TOTAL_FONT
    due_row = row_at
    # which cell it came from - the two voucher forms keep it in different
    # places, so the summary says where it read rather than leaving it implied
    if result.get("due_cell"):
        note = ws.cell(row_at, 4, f"read from {result['due_cell']}")
        note.font = Font(size=9, italic=True, color="FF808080")

    # Section B labor and the cash advance, both written unconditionally so the
    # row offsets are stable for the tracker to reference. Blank advance is
    # written as blank, not zero - the distinction is real on the voucher.
    row_at += 1
    ws.cell(row_at, 1, "of which Section B labor").font = LABEL_FONT
    labor = ws.cell(row_at, 3, result.get("labor"))
    labor.number_format = FMT["money"]
    if result.get("labor"):
        labor.fill = WARN_FILL
    labor_row = row_at

    row_at += 1
    ws.cell(row_at, 1, "Cash advance / prepaid").font = LABEL_FONT
    advance = ws.cell(row_at, 3, result.get("advance"))
    advance.number_format = FMT["money"]
    if result.get("advance"):
        advance.fill = WARN_FILL
    advance_row = row_at

    row_at += 1
    ws.cell(row_at, 1, "Voucher Says (D39)").font = LABEL_FONT
    said = ws.cell(row_at, 3, result["reported"].get("mileage"))
    said.number_format = FMT["money"]
    row_at += 1
    ws.cell(row_at, 1, "Lines Add To").font = LABEL_FONT
    ws.cell(row_at, 3, f"=C{total_row}").number_format = FMT["money"]

    if result["notes"]:
        row_at += 2
        ws.cell(row_at, 1, "Notes").font = LABEL_FONT
        for note in result["notes"]:
            row_at += 1
            ws.cell(row_at, 1, note)

    autosize(ws, [12, 11, 12, 12, 12, 12, 12, 16, 18, 11, 12, 10, 26, 34])
    ws.freeze_panes = ws.cell(first_data, 1)
    return {"title": title, "total_row": total_row, "due_row": due_row,
            "labor_row": labor_row, "advance_row": advance_row}


MASTER_COLUMNS = (
    [("Employee", 24, None), ("Voucher Date", 13, "date"),
     ("Trip Start", 12, "date"), ("Trip End", 12, "date"),
     ("Lines", 7, "miles"), ("POV Miles", 11, "miles")]
    + [(CATEGORY_LABELS[k], 15, "money") for k in CATEGORIES]
    + [("Due Employee", 15, "money"), ("Rates Used", 20, None), ("Rate Flags", 10, "miles"),
       ("Status", 9, None), ("Source File", 38, None), ("Notes", 60, None)]
)


def ref(title, cell):
    """Cross-sheet reference, safe for spaces and apostrophes in the name."""
    return f"='{title.replace(chr(39), chr(39) * 2)}'!{cell}"


def write_master_sheet(ws, results, sheet_info, expected_rate=DEFAULT_RATE):
    ws["A1"] = "Voucher Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"{len(results)} voucher(s) parsed {dt.date.today().isoformat()}"
                f"  ·  mileage checked against {expected_rate}/mile")
    ws["A2"].font = Font(size=9, italic=True)

    top = 4
    for idx, (heading, _, _) in enumerate(MASTER_COLUMNS, start=1):
        cell = ws.cell(top, idx, heading)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    row_at = top
    for result in results:
        row_at += 1
        dates = [ln["date"] for ln in result["lines"] if ln.get("date")]
        values = [
            result["employee"] or "(missing)",
            result["voucher_date"],
            min(dates) if dates else None,
            max(dates) if dates else None,
            len(result["lines"]),
            result["computed"].get("pov_miles", 0.0),
        ]
        values += [result["computed"].get(k, 0.0) for k in CATEGORIES]
        rates = sorted(set(result.get("rates_used", [])))
        values += [
            result["due_employee"],
            ", ".join(f"{r:g}" for r in rates),
            result.get("rate_issues", 0),
            result["status"],
            result["source"],
            "; ".join(result["notes"]),
        ]
        for idx, ((_, _, kind), value) in enumerate(zip(MASTER_COLUMNS, values), start=1):
            cell = ws.cell(row_at, idx, value)
            cell.border = BORDER
            if kind:
                cell.number_format = FMT[kind]

        # point the money columns at the voucher sheets so the Master is live
        info = sheet_info.get(result["source"])
        if info:
            for offset in range(8):  # POV Miles + the seven categories
                src = f"{get_column_letter(2 + offset)}{info['total_row']}"
                ws.cell(row_at, 6 + offset, ref(info["title"], src))
            ws.cell(row_at, 14, ref(info["title"], f"C{info['due_row']}"))

        if result.get("rate_issues"):
            ws.cell(row_at, 16).fill = WARN_FILL
        status_cell = ws.cell(row_at, len(MASTER_COLUMNS) - 2)
        if result["status"] == "WARN":
            status_cell.fill = WARN_FILL
        elif result["status"] == "ERROR":
            status_cell.fill = ERR_FILL

        info = sheet_info.get(result["source"])
        if info:
            link = ws.cell(row_at, 1)
            link.hyperlink = f"#'{info['title']}'!A1"
            link.font = Font(color="FF0563C1", underline="single")

    last_data = row_at

    # grand total - live, so it follows any correction made on a voucher sheet
    row_at += 1
    label = ws.cell(row_at, 1, "ALL VOUCHERS")
    label.font, label.fill, label.border = TOTAL_FONT, TOTAL_FILL, BORDER
    for idx in range(2, len(MASTER_COLUMNS) + 1):
        cell = ws.cell(row_at, idx)
        cell.fill, cell.font, cell.border = TOTAL_FILL, TOTAL_FONT, BORDER
    for idx in [5, 16] + list(range(6, 15)):
        col = get_column_letter(idx)
        cell = ws.cell(row_at, idx, f"=SUM({col}{top + 1}:{col}{last_data})")
        cell.number_format = FMT["money"] if 7 <= idx <= 14 else FMT["miles"]

    autosize(ws, [w for _, w, _ in MASTER_COLUMNS])
    ws.freeze_panes = ws.cell(top + 1, 1)
    ws.auto_filter.ref = f"A{top}:{get_column_letter(len(MASTER_COLUMNS))}{row_at - 1}"
    return ws


# ----------------------------------------------------------------------- main

def next_version(folder: Path, stem: str) -> Path:
    """Never overwrite a summary. Find the highest -vNN already in the folder
    and return the next one, so every run leaves an auditable trail."""
    highest = 0
    pattern = re.compile(re.escape(stem) + r"-v(\d+)\.xlsx$", re.IGNORECASE)
    if folder.is_dir():
        for existing in folder.glob(f"{stem}-v*.xlsx"):
            found = pattern.search(existing.name)
            if found:
                highest = max(highest, int(found.group(1)))
    return folder / f"{stem}-v{highest + 1:02d}.xlsx"


def main(argv=None):
    ap = argparse.ArgumentParser(description="Parse West Run, LLC travel vouchers.")
    ap.add_argument("--project", help="a 04-Projects/<Project> folder; implies its stage folders")
    ap.add_argument("--month", help=argparse.SUPPRESS)  # deprecated alias for --project
    ap.add_argument("--inbox", help="folder of voucher .xlsx files")
    ap.add_argument("--out", help="output workbook path")
    ap.add_argument("--exceptions", help="folder for unreadable vouchers")
    ap.add_argument("--rate", type=float, default=DEFAULT_RATE,
                    help=f"mileage rate the contractors should have billed (default {DEFAULT_RATE})")
    ap.add_argument("--bill-rate", type=float, default=None,
                    help="mileage rate West Run BILLS at on the tracker, if it differs "
                         "from --rate. Defaults to --rate, which is the normal case")
    args = ap.parse_args(argv)

    # --month was the flag before the cycle moved from months to projects on
    # 2026-08-09. Still accepted so an old launcher does not just fail.
    folder = args.project or args.month
    if args.month and not args.project:
        print("note: --month is deprecated, use --project")

    if folder:
        # resolve() first: a session started inside the project folder passes
        # ".", and Path(".").name is an empty string - which silently produced
        # "Voucher-Summary--v01.xlsx" with the project name missing.
        project = Path(folder).resolve()
        inbox = Path(args.inbox) if args.inbox else stage_dir(project, "01", "01-Inbox")
        # The workbook name carries the exercise, which is not known until the
        # vouchers have been read - so it is decided after parsing, not here.
        out = Path(args.out) if args.out else None
        exceptions = Path(args.exceptions) if args.exceptions else None
    else:
        if not args.inbox or not args.out:
            ap.error("give --project, or both --inbox and --out")
        inbox, out = Path(args.inbox), Path(args.out)
        exceptions = Path(args.exceptions) if args.exceptions else None

    if not inbox.is_dir():
        print(f"inbox not found: {inbox}", file=sys.stderr)
        return 2

    submissions = collect_submissions(inbox)
    files = [voucher for voucher, _r, _n in submissions if voucher]
    receipts_by_voucher = {v: r for v, r, _n in submissions if v}
    orphans = [(r, n) for v, r, n in submissions if not v]
    if not files and not orphans:
        print(f"no vouchers in {inbox}")
        return 1

    print(f"reading {len(files)} voucher(s) from {inbox}")
    print(f"mileage rate checked against {args.rate}/mile")
    results = []
    for path in files:
        result = parse_voucher(path, args.rate)
        found = receipts_by_voucher.get(path, [])
        result["receipts"] = [p.name for p in found]
        if not found:
            result["notes"].append("no receipts in this person's folder")
            if result["status"] == "OK":
                result["status"] = "WARN"
        results.append(result)
    results.extend(orphan_result(r, n) for r, n in orphans)

    # Alphabetical by name, not by filename. Sheet order follows creation order
    # and the tracker rows follow this list, so sorting once here puts the tabs
    # and the front page in the same order the roster is read in.
    results.sort(key=name_sort_key)

    if out is None:
        # Exercise name off the vouchers (C14) rather than the folder name - it
        # is what the client calls the job, and it travels on the file.
        exercise = next((r["exercise"] for r in results if r.get("exercise")), "")
        out = tracker_name(stage_dir(project, "03", "03-WR Invoice Tracker"),
                           exercise or project.name)

    wb = Workbook()
    tracker = wb.active
    tracker.title = "Invoice Tracker"

    # Per-voucher sheets first: the front page references them, so they have to
    # exist before it can be written.
    used, sheet_info = set(), {}
    for result in results:
        title = sheet_title(result["employee"], result["source"], used)
        sheet_info[result["source"]] = write_voucher_sheet(wb, result, title)

    # One rate answers both questions unless told otherwise: the rate West Run
    # bills at is the rate contractors are supposed to have billed at.
    bill_rate = args.bill_rate if args.bill_rate is not None else args.rate

    totals_row = write_tracker(tracker, results, sheet_info, args.rate, bill_rate)

    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out)
    except PermissionError:
        # almost always the previous summary still open in Excel
        stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
        alt = out.with_name(f"{out.stem}-{stamp}{out.suffix}")
        wb.save(alt)
        print(f"\n!! {out.name} is open in Excel - could not overwrite it.")
        print(f"!! Wrote {alt.name} instead. Close the old one and rerun to replace it.")
        out = alt

    # unreadable vouchers get copied out with a reason, originals left alone
    broken = [r for r in results if r["status"] == "ERROR"]
    if broken and exceptions:
        exceptions.mkdir(parents=True, exist_ok=True)
        with open(exceptions / "_exceptions.csv", "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow(["source_file", "employee", "reason"])
            for result in broken:
                # The summary is already on disk at this point. A copy that
                # fails - long path, file locked, permissions - must not take
                # the whole run down with it, or the launcher reports "nothing
                # was written" over a workbook that was written.
                try:
                    shutil.copy2(inbox / result["source"], exceptions / result["source"])
                except OSError as exc:
                    print(f"!! could not copy {result['source']} to {exceptions}: {exc}")
                writer.writerow([result["source"], result["employee"], "; ".join(result["notes"])])

    for result in results:
        flag = {"OK": "  ok  ", "WARN": " warn ", "ERROR": "ERROR "}[result["status"]]
        print(f"  [{flag}] {result['employee'] or '(no name)':<22} {len(result['lines']):>3} lines  {result['source']}")
        for note in result["notes"]:
            print(f"           - {note}")

    counts = {s: sum(1 for r in results if r["status"] == s) for s in ("OK", "WARN", "ERROR")}
    print(f"\n{counts['OK']} clean, {counts['WARN']} with warnings, {counts['ERROR']} failed")
    print(f"written: {out}")

    # Stage 05-Final Output. Only categories this project actually spent in get
    # a folder, so an empty one means a receipt is missing rather than that the
    # category was never used.
    if folder:
        final, used, undecidable, copied = scaffold_final_output(
            Path(folder), results, files)
        print(f"staged:  {final}")
        print(f"         {copied} voucher(s) copied to Vouchers/")
        print(f"         expect receipts in: {', '.join(used) or 'none'}")
        if undecidable:
            print(f"         plus, if keyed by hand: {', '.join(undecidable)}")
        print("         category folders are created by Step 4, only where a "
              "receipt is filed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
