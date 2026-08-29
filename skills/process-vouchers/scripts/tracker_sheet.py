#!/usr/bin/env python3
# ---------------------------------------------------------------------
# COPY - do not edit here. The original is 02-Tool/tracker_sheet.py in the LSG-005
# project folder; this file is overwritten by 05-Plugin/build_plugin.py.
# ---------------------------------------------------------------------
"""
LSG-005 — the WR Invoice Tracker front page, written from parsed vouchers.

Replaces the flat Master sheet. The output workbook is now shaped like the
tracker West Run already keeps (01-Reference/WR Invoice Tracker (Example).xlsx),
so a run drops into their existing workflow instead of asking them to learn a
new one.

Two kinds of sheet:

  Invoice Tracker    the front page - one row per independent contractor
  <Lastname F>       one sheet per voucher, the line detail and mileage audit

Every money cell the front page can get from a voucher is a LIVE CROSS-SHEET
FORMULA into that person's own sheet. Correct a line on their sheet and the
front page follows. That is the "individual pages push to the front page"
behaviour.

Columns a voucher cannot know - West Run's own direct costs, labor rates and
days, MISC - are left empty and shaded, so what still needs
keying is visible at a glance rather than discovered at invoice time.

Two columns added that West Run's sheet does not have:

  Cash Advance        The voucher's I44. Their tracker has no advance column at
                      all, so it could never reconcile to a voucher that carried
                      one - the voucher's Due is always NET of the advance while
                      their Total To Be Paid To IC is gross. Subtracted in
                      to_be_paid, which is what makes Variance mean something.
  Labor per Voucher   The Section B labor the contractor claimed, alongside the
                      rate x days West Run keys in. When the labor columns have
                      not been filled yet, Variance equals this number exactly,
                      so the sheet tells you what is missing instead of just
                      that something is.

Deliberately NOT reproduced from the original:
  * the hardcoded 0.725 in seventeen separate mileage formulas. One editable
    rate cell (D3) drives every row.
  * the hand-patched constants - '+12' in H4/H6, '+0.01' in U7, '-0.01' in U10.
    Those exist to force a row to agree with a voucher. Here the voucher's own
    figure (AK) and the computed figure (AL) both stand and the gap between
    them is a Variance column. The disagreement is the finding.
  * the margin block entirely - the rate card, CDI invoice, debts, delta,
    markup and profit margin. Dropped 2026-08-11. It is West Run's commercial
    position, this workbook is a reimbursement record, and the safest place for
    markup is not in a file that gets forwarded.
"""

from __future__ import annotations

import datetime as dt
import os
import re
import shutil
import sys
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ Lark brand
# Read out of 02-Brand/Logos/*-green.svg, not guessed.
LARK_GREEN = "FF304945"
LARK_GREEN_DEEP = "FF2F4A43"
LARK_TINT = "FFE7EDEB"      # header block / totals wash
LARK_RULE = "FFB9C6C2"
NEEDS_ENTRY = "FFFDF3E2"    # columns a voucher cannot fill
VARIANCE = "FFF8CBAD"
WARN = "FFFFF2CC"
INK = "FF1B2B28"

HDR_FONT = Font(bold=True, color="FFFFFFFF", size=9, name="Calibri")
HDR_FILL = PatternFill("solid", start_color=LARK_GREEN)
TITLE_FONT = Font(bold=True, size=16, color=LARK_GREEN, name="Calibri")
SUB_FONT = Font(size=9, italic=True, color="FF5A6B67")
LABEL_FONT = Font(bold=True, size=10, color=INK)
TOTAL_FONT = Font(bold=True, size=10, color=INK)
TOTAL_FILL = PatternFill("solid", start_color=LARK_TINT)
THIN = Side(style="thin", color=LARK_RULE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def long_path(path):
    r"""Windows refuses paths over 260 characters unless they carry the \\?\
    prefix. Stage folder names plus the filename conventions cross that once a
    project sits a few levels deep in OneDrive, and it fails per-file depending
    on how long that one's fields happen to be - so it looks random rather than
    structural. Every filesystem write goes through this."""
    text = str(Path(path).resolve())
    if sys.platform == "win32" and not text.startswith("\\\\?\\"):
        return "\\\\?\\" + text
    return text


def split_name(full):
    """'Larkin, Brandon, R.' and 'SANDS JR, NEAL F' -> (last, first, middle)

    Two conventions in play. The older form comma-separates the middle initial;
    the West Run form separates it with a space, so a naive split leaves
    'NEAL F' sitting in the forename. A trailing single letter is an initial,
    never a name - 'MARY ANN' keeps both words."""
    if not full:
        return "", "", ""
    parts = [p.strip() for p in str(full).split(",")]
    last = parts[0] if parts else ""
    first = parts[1] if len(parts) > 1 else ""
    middle = parts[2].rstrip(".") if len(parts) > 2 else ""
    if not middle:
        found = re.match(r"^(.*?)\s+([A-Za-z])\.?$", first)
        if found:
            first, middle = found.group(1).strip(), found.group(2)
    return last, first, middle


def initial_surname(employee):
    """'SANDS JR, NEAL F' -> 'N. Sands Jr'. Blank if there is no name."""
    last, first, _mid = split_name(employee)
    if not last:
        return ""
    return f"{first[:1].upper()}. {last.title()}" if first else last.title()


def safe_field(text, fallback="UNKNOWN"):
    """One filename field, legal on Windows. Never silently empty."""
    text = re.sub(r'[\\/:*?"<>|]', "-", str(text or "")).strip()
    text = re.sub(r"\s+-\s+", " ", text)          # " - " is the field separator
    return re.sub(r"\s{2,}", " ", text).strip(" .") or fallback


def voucher_name(result, path, project):
    """YYYY.MM.DD - Exercise - F. Lastname.ext

    Date is the voucher's own date (`I9`), the nearest thing it has to "when
    this document was made" - the same rule receipts follow. Falls back to the
    last day of travel, then to an obviously wrong stamp rather than today's
    date, which would silently change on every re-run.

    Project is the exercise name off the voucher (`C14`) rather than the folder
    name: it is what the client calls the job, and it is carried on the document
    itself. Falls back to the folder name."""
    result = result or {}
    date = result.get("voucher_date")
    if not date:
        dates = [ln["date"] for ln in result.get("lines", []) if ln.get("date")]
        date = max(dates) if dates else None
    stamp = date.strftime("%Y.%m.%d") if date else "0000.00.00"
    exercise = safe_field(result.get("exercise") or project.name, project.name)
    who = safe_field(initial_surname(result.get("employee", "")), "UNKNOWN")
    return f"{stamp} - {exercise} - {who}{path.suffix.lower()}"


def tracker_name(folder, project_name, when=None):
    """'2026.08.11 - Invoice Tracker - NSW 2607-A.xlsx'

    The date is the RUN date. Unlike a voucher or a receipt the tracker has no
    document date of its own - it is generated, and it is generated today.

    A second run the same day gets ' (2)', ' (3)' rather than overwriting. That
    matters more than it looks: re-running as vouchers trickle in is the normal
    way this pipeline is used, and a tracker somebody has already worked from
    must not be silently replaced underneath them."""
    when = when or dt.date.today()
    base = (f"{when.strftime('%Y.%m.%d')} - Invoice Tracker - "
            f"{safe_field(project_name, 'Project')}")
    folder.mkdir(parents=True, exist_ok=True)
    target = folder / f"{base}.xlsx"
    count = 2
    while os.path.exists(long_path(target)):
        target = folder / f"{base} ({count}).xlsx"
        count += 1
    return target


# The flat layout, introduced 2026-08-28 when the five manual stages collapsed
# into "put things in, get things out". Only two folders are anybody's business:
#
#   Inbox/    one folder per person, their voucher and their receipts
#   Output/   what leaves - vouchers, receipts by category, the audit workbook,
#             the QuickBooks import file
#
# _work/ is the pipeline's own scratch: the un-annotated tracker, _extract.json,
# _matches.json, the run log. Named with an underscore so it sorts out of the
# way and reads as "not yours". Deleting it loses nothing that Output does not
# already carry.
STAGE_ALIASES = {
    "01": ("Inbox",),
    "02": ("_work",),
    "03": ("_work",),
    "04": ("_work",),
    "05": ("Output",),
}


def stage_dir(project, prefix, fallback=None):
    """Find a stage folder, by leading number or by flat name.

    Three layouts have to resolve, because a project folder created under an
    older one must keep working after the tool is updated underneath it:

      1. Numbered stages - '01-Inbox (Manually Add People)'. The tail is a human
         instruction that gets reworded, so the NUMBER is what is matched. Any
         project created before 2026-08-28 looks like this.
      2. Flat - 'Inbox' and 'Output'. The current shape.
      3. Neither, because nothing has been created yet - resolve to the flat
         name so the caller creates the current layout, never the old one.

    Numbered wins when both exist. That is deliberate: a half-migrated folder
    should keep behaving exactly as it did rather than silently splitting its
    output across two structures."""
    project = Path(project)
    aliases = STAGE_ALIASES.get(prefix, ())
    if project.is_dir():
        for child in sorted(project.iterdir()):
            if child.is_dir() and child.name.startswith(f"{prefix}-"):
                return child
        for alias in aliases:
            child = project / alias
            if child.is_dir():
                return child
    if aliases:
        return project / aliases[0]
    return project / fallback


MONEY = '#,##0.00'
MILES = '#,##0'
DATE = 'yyyy-mm-dd'
RATE = '0.000'

# Where the header block lives. The mileage rate is a single cell so it can be
# changed once and audited, instead of living inside every formula.
CELL_TDY, CELL_EXERCISE, CELL_RATE = "D1", "D2", "D3"
HEADER_ROW = 5
FIRST_DATA = HEADER_ROW + 1

# On each per-voucher sheet the TOTAL row runs: A date, B POV miles,
# C mileage $, D airfare, E car rental, F fuel, G hotel, H per diem, I parking.
SRC = {"pov": "B", "mileage": "C", "airfare": "D", "car": "E",
       "fuel": "F", "lodging": "G", "per_diem": "H", "parking": "I"}

# key, header, width, format, kind
#   kind: "voucher" filled from the voucher sheet · "formula" computed here
#         "manual" West Run's own cost, left for a human · "meta" parser output
COLUMNS = [
    ("name",        "Name",                        24, None,  "voucher"),
    ("first_day",   "First Travel Day",            13, DATE,  "voucher"),
    ("pd_first",    "Travel Per Diem",             12, MONEY, "voucher"),
    ("pd_days",     "Full Per Diem Days",          11, MILES, "voucher"),
    ("pd_rate",     "Full Per Diem",               11, MONEY, "voucher"),
    ("last_day",    "Last Day of Travel",          13, DATE,  "voucher"),
    ("pd_last",     "Travel Per Diem2",            12, MONEY, "voucher"),
    ("pd_total",    "Per Diem Total",              13, MONEY, "voucher"),
    ("air_to",      "Airfare To",                  11, MONEY, "manual"),
    ("air_from",    "Airfare From",                11, MONEY, "manual"),
    ("air_total",   "Total Airfare",               12, MONEY, "voucher"),
    ("baggage",     "Baggage",                     10, MONEY, "manual"),
    ("air_ic",      "Airfare Paid to IC",          14, MONEY, "formula"),
    ("lodging",     "Lodging",                     12, MONEY, "manual"),
    ("lodging_ic",  "Lodging Paid to IC",          14, MONEY, "voucher"),
    ("vehicle",     "Rental Vehicle",              13, MONEY, "manual"),
    ("vehicle_ic",  "Rental Vehicle Paid to IC",   16, MONEY, "voucher"),
    ("rfuel",       "Rental Fuel",                 11, MONEY, "manual"),
    ("rfuel_ic",    "Fuel Paid to IC",             13, MONEY, "voucher"),
    ("pov",         "POV Miles",                   10, MILES, "voucher"),
    ("mileage",     "Mileage Billed",              13, MONEY, "formula"),
    ("parking",     "Parking/Tolls/Taxis",         14, MONEY, "manual"),
    ("parking_ic",  "Parking Paid to IC",          14, MONEY, "voucher"),
    ("travel_tot",  "Travel Total",                13, MONEY, "formula"),
    ("paid_ic",     "Total Travel Paid to ICs",    16, MONEY, "formula"),
    ("paid_wr",     "Total Travel Paid by WR",     16, MONEY, "formula"),
    ("misc",        "MISC.",                       10, MONEY, "manual"),
    ("misc_ic",     "MISC. Paid to IC",            13, MONEY, "manual"),
    ("labor_rate",  "Labor Rate",                  11, MONEY, "manual"),
    ("labor_oh",    "Labor Rate w/ P/OH",          14, MONEY, "manual"),
    ("half_days",   "1/2 Labor Days",              11, MILES, "manual"),
    ("full_days",   "Full Labor Days",             12, MILES, "manual"),
    ("labor_tot",   "Labor Total",                 12, MONEY, "formula"),
    ("labor_tot_oh", "Labor Total w/ P/OH",        14, MONEY, "formula"),
    # Not on West Run's sheet - see "Two columns added" in the module docstring.
    ("labor_voucher", "Labor per Voucher",         14, MONEY, "voucher"),
    ("advance",     "Cash Advance",                12, MONEY, "voucher"),
    ("submitted",   "IC Invoice as submitted",     16, MONEY, "voucher"),
    ("to_be_paid",  "Total To Be Paid To IC",      16, MONEY, "formula"),
    ("variance",    "Variance",                    11, MONEY, "formula"),
    ("receipts",    "Receipts",                     9, MILES, "meta"),
    ("status",      "Status",                       9, None,  "meta"),
    ("notes",       "Notes",                       64, None,  "meta"),
]

IDX = {key: n for n, (key, *_rest) in enumerate(COLUMNS, start=1)}


def col(key):
    return get_column_letter(IDX[key])


def formula_for(key, row):
    """The computed columns, expressed exactly as West Run's own sheet does
    them - except the mileage rate, which now points at the rate cell."""
    c = {k: col(k) for k in IDX}
    r = row
    return {
        "air_ic":       f"={c['air_total']}{r}",
        "mileage":      f"={c['pov']}{r}*${CELL_RATE[0]}${CELL_RATE[1:]}",
        "travel_tot":   (f"={c['pd_total']}{r}+{c['air_total']}{r}+{c['lodging']}{r}"
                         f"+{c['vehicle']}{r}+{c['rfuel']}{r}+{c['mileage']}{r}"
                         f"+{c['parking']}{r}+{c['baggage']}{r}"),
        "paid_ic":      f"={c['travel_tot']}{r}-{c['paid_wr']}{r}",
        "paid_wr":      (f"={c['travel_tot']}{r}-{c['mileage']}{r}-{c['parking']}{r}"
                         f"-{c['rfuel_ic']}{r}-{c['vehicle_ic']}{r}-{c['lodging_ic']}{r}"
                         f"-{c['air_ic']}{r}-{c['pd_total']}{r}"),
        "labor_tot":    (f"=(({c['labor_rate']}{r}/2)*{c['half_days']}{r})"
                         f"+({c['labor_rate']}{r}*{c['full_days']}{r})"),
        "labor_tot_oh": (f"=(({c['labor_oh']}{r}/2)*{c['half_days']}{r})"
                         f"+({c['labor_oh']}{r}*{c['full_days']}{r})"),
        # Gross reimbursement LESS any cash advance already paid. West Run's
        # sheet has no advance column, so it could never tie to a voucher that
        # carried one; the voucher's Due is always net of it.
        "to_be_paid":   (f"={c['pd_total']}{r}+{c['air_ic']}{r}+{c['lodging_ic']}{r}"
                         f"+{c['vehicle_ic']}{r}+{c['rfuel_ic']}{r}+{c['mileage']}{r}"
                         f"+{c['parking_ic']}{r}+{c['labor_tot']}{r}+{c['misc_ic']}{r}"
                         f"-{c['advance']}{r}"),
        # The whole point: what the contractor billed against what the sheet
        # says they are owed. Non-zero means somebody has to look.
        "variance":     f"=ROUND({c['submitted']}{r}-{c['to_be_paid']}{r},2)",
    }.get(key)


def sheet_ref(title, cell):
    """Cross-sheet reference, safe for spaces and apostrophes in the name."""
    return f"'{title.replace(chr(39), chr(39) * 2)}'!{cell}"


def per_diem_shape(result):
    """Split the voucher's per-diem column into the tracker's four fields:
    first travel day, full days, the full-day rate, last travel day.

    The tracker treats per diem as rate x days plus two partial travel days.
    Real vouchers do not always fit that, which is why the original sheet has
    '+12' hand-typed into two of its formulas. This returns the best fit and
    the leftover, and the caller reports the leftover instead of hiding it."""
    values = [ln.get("per_diem") for ln in result["lines"]]
    values = [v for v in values if v is not None]
    if not values:
        return None, 0, None, None, 0.0
    first, last = values[0], values[-1]
    middle = values[1:-1] if len(values) > 2 else []
    rate = max(set(middle), key=middle.count) if middle else None
    days = len(middle)
    fitted = first + (last if len(values) > 1 else 0) + (rate or 0) * days
    total = round(sum(values), 2)
    return first, days, rate, (last if len(values) > 1 else None), round(total - fitted, 2)


def write_tracker(ws, results, sheet_info, expected_rate, mileage_rate=0.725):
    """The front page. One row per contractor, live into their own sheet."""
    project_tdy = next((r["tdy"] for r in results if r.get("tdy")), "")
    exercise = next((r["exercise"] for r in results if r.get("exercise")), "")

    ws["A1"] = "TDY Destination:"
    ws[CELL_TDY] = project_tdy
    ws["A2"] = "Purpose of travel:"
    ws[CELL_EXERCISE] = exercise
    ws["A3"] = "Mileage rate:"
    ws[CELL_RATE] = mileage_rate
    ws[CELL_RATE].number_format = RATE
    ws[CELL_RATE].fill = PatternFill("solid", start_color=NEEDS_ENTRY)
    ws[CELL_RATE].font = Font(bold=True, size=10, color=INK)
    for cell in ("A1", "A2", "A3"):
        ws[cell].font = LABEL_FONT
    for cell in (CELL_TDY, CELL_EXERCISE):
        ws[cell].font = Font(bold=True, size=11, color=LARK_GREEN_DEEP)

    ws["G1"] = "WR INVOICE TRACKER"
    ws["G1"].font = TITLE_FONT
    ws["G2"] = (f"{len(results)} voucher(s) parsed {dt.date.today().isoformat()}"
                f"  ·  mileage audited against {expected_rate}/mile"
                f"  ·  shaded columns still need keying")
    ws["G2"].font = SUB_FONT

    for idx, (_key, heading, _w, _f, kind) in enumerate(COLUMNS, start=1):
        cell = ws.cell(HEADER_ROW, idx, heading)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        if kind == "manual":
            cell.fill = PatternFill("solid", start_color=LARK_GREEN_DEEP)

    row_at = HEADER_ROW
    for result in results:
        row_at += 1
        info = sheet_info.get(result["source"])
        title = info["title"] if info else None
        trow = info["total_row"] if info else None
        drow = info["due_row"] if info else None

        first_pd, pd_days, pd_rate, last_pd, leftover = per_diem_shape(result)
        dates = [ln["date"] for ln in result["lines"] if ln.get("date")]

        values = {
            "name": result["employee"] or "(missing)",
            "first_day": min(dates) if dates else None,
            "pd_first": first_pd,
            "pd_days": pd_days,
            "pd_rate": pd_rate,
            "last_day": max(dates) if dates else None,
            "pd_last": last_pd,
            "receipts": len(result.get("receipts") or []),
            "status": result["status"],
            "notes": "; ".join(result["notes"]),
        }

        # money that comes off the voucher, as live references into its sheet
        if title and trow:
            values["pd_total"] = f"={sheet_ref(title, SRC['per_diem'] + str(trow))}"
            values["air_total"] = f"={sheet_ref(title, SRC['airfare'] + str(trow))}"
            values["lodging_ic"] = f"={sheet_ref(title, SRC['lodging'] + str(trow))}"
            values["vehicle_ic"] = f"={sheet_ref(title, SRC['car'] + str(trow))}"
            values["rfuel_ic"] = f"={sheet_ref(title, SRC['fuel'] + str(trow))}"
            values["pov"] = f"={sheet_ref(title, SRC['pov'] + str(trow))}"
            values["parking_ic"] = f"={sheet_ref(title, SRC['parking'] + str(trow))}"
        if title and drow:
            values["submitted"] = f"={sheet_ref(title, 'C' + str(drow))}"
        if title and info.get("labor_row"):
            values["labor_voucher"] = f"=N({sheet_ref(title, 'C' + str(info['labor_row']))})"
        if title and info.get("advance_row"):
            values["advance"] = f"=N({sheet_ref(title, 'C' + str(info['advance_row']))})"

        for idx, (key, _h, _w, fmt, kind) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row_at, idx)
            if kind == "formula":
                cell.value = formula_for(key, row_at)
            elif key in values:
                cell.value = values[key]
            cell.number_format = fmt or 'General'
            cell.border = BORDER
            if kind == "manual":
                cell.fill = PatternFill("solid", start_color=NEEDS_ENTRY)

        ws.cell(row_at, IDX["variance"]).fill = PatternFill("solid", start_color=VARIANCE)
        if not result.get("receipts"):
            ws.cell(row_at, IDX["receipts"]).fill = PatternFill("solid", start_color=VARIANCE)
        if result["status"] in ("WARN", "ERROR"):
            ws.cell(row_at, IDX["status"]).fill = PatternFill("solid", start_color=WARN)

        # per diem that will not fit rate x days - the reason the original
        # sheet has '+12' typed into two formulas
        if leftover and abs(leftover) > 0.005:
            note = ws.cell(row_at, IDX["notes"])
            note.value = ((note.value + "; ") if note.value else "") + (
                f"per diem lines exceed rate x days by {leftover:.2f} - "
                f"Per Diem Total is the voucher's own figure")
            note.fill = PatternFill("solid", start_color=WARN)

        if title:
            link = ws.cell(row_at, 1)
            link.hyperlink = f"#'{title}'!A1"
            link.font = Font(color="FF0563C1", underline="single")

    last_ic = last_data = row_at

    row_at += 1
    label = ws.cell(row_at, 1, "TOTALS")
    label.font, label.fill, label.border = TOTAL_FONT, TOTAL_FILL, BORDER
    for idx, (_key, _h, _w, fmt, _kind) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row_at, idx)
        cell.fill, cell.font, cell.border = TOTAL_FILL, TOTAL_FONT, BORDER
        if idx == 1 or fmt in (None, DATE):
            continue
        letter = get_column_letter(idx)
        # SUBTOTAL across the whole range, every column. Their sheet mixes
        # SUBTOTAL with plain SUM, and K/U stop at row 20 - so anything keyed
        # into the mission-fuel rows for Total Airfare or Mileage Billed is
        # silently left out of their totals.
        cell.value = f"=SUBTOTAL(109,{letter}{FIRST_DATA}:{letter}{last_data})"
        cell.number_format = fmt
    totals_row = row_at

    # Total Flights - I30/J30 on their sheet: airfare out plus airfare back.
    row_at += 1
    ws.cell(row_at, IDX["air_to"], "Total Flights:").font = LABEL_FONT
    flights = ws.cell(row_at, IDX["air_from"],
                      f"={col('air_to')}{totals_row}+{col('air_from')}{totals_row}")
    flights.number_format, flights.font = MONEY, TOTAL_FONT

    for idx, (_k, _h, width, *_r) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[HEADER_ROW].height = 30
    ws.freeze_panes = ws.cell(FIRST_DATA, 2)
    # Filter the contractor rows only - sorting the mission-fuel block into
    # them would put entries under the wrong heading.
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLUMNS))}{last_ic}"
    return totals_row


# ------------------------------------------------- final output scaffolding

# The receipt categories are the tracker's own expense columns. Per Diem, POV
# Mileage and Labor are deliberately absent: they are computed from a rate, not
# evidenced by a receipt, so a folder for them would sit empty and read as a
# gap. Folder names avoid '/' and a trailing '.', neither of which survives on
# Windows.
#
# key = the computed total the parser reads off a voucher, or None where the
# column is keyed by hand and so cannot be judged at parse time.
RECEIPT_CATEGORIES = [
    ("Airfare",                "airfare"),
    ("Lodging",                "hotel"),
    ("Rental Vehicle",         "car_rental"),
    ("Rental Fuel",            "fuel"),
    ("Parking, Tolls & Taxis", "taxi"),
    ("Baggage",                None),
    ("MISC",                   None),
]


def used_categories(results):
    """Which categories this project actually spent money in.

    Advisory only - nothing is created from this. It tells whoever runs Step 4
    which categories to expect receipts in, so a category with spend and no
    receipt is noticed.

    Baggage and MISC are hand-keyed columns with nothing behind them on a
    voucher, so they cannot be decided here and come back separately."""
    used, undecidable = [], []
    for label, key in RECEIPT_CATEGORIES:
        if key is None:
            undecidable.append(label)
            continue
        total = sum((r.get("computed") or {}).get(key) or 0 for r in results)
        if round(total, 2) > 0:
            used.append(label)
    return used, undecidable


def scaffold_final_output(project, results, vouchers):
    """Create 05-Final Output for this project and stage what already exists.

    Vouchers/   copies of the vouchers exactly as they arrived
    (root)      the checked tracker, placed by Step 4

    **No category folders are created here.** Step 4 creates each one at the
    moment it files the first receipt into it, so a folder on disk is evidence
    that a receipt exists. Pre-creating them from voucher spend looked helpful
    but inverted the signal: an empty `Rental Vehicle` folder read as a missing
    receipt rather than as a category nobody used.

    The category list is still returned, as a heads-up about which categories
    receipts are expected in - reported, not built.

    Vouchers are COPIED, never moved. 01-Inbox stays the untouched record of
    what was received."""
    project = Path(project).resolve()
    final = stage_dir(project, "05", "05-Final Output")
    vouchers_dir = final / "Vouchers"
    used, undecidable = used_categories(results)
    os.makedirs(long_path(vouchers_dir), exist_ok=True)

    # Renamed on the way in. What a contractor called their file is noise -
    # 'Voucher (3).xlsx' says nothing - so the copy carries the date, the
    # exercise and whose it is. The original keeps its name in 01-Inbox.
    by_source = {r.get("source"): r for r in results}
    copied = 0
    for path in vouchers:
        target = vouchers_dir / voucher_name(by_source.get(path.name), path, project)
        shutil.copy2(long_path(path), long_path(target))
        copied += 1
    return final, used, undecidable, copied
